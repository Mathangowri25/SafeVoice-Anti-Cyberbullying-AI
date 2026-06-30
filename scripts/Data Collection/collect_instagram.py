"""
Instagram Comment Scraper using Selenium (browser automation)
Uses robust XPath/aria selectors instead of Instagram's changing CSS class names.

Install:
    pip install selenium webdriver-manager pandas openpyxl python-dotenv

Usage:
    python scrape_instagram.py
"""

import os
import re
import time
import random
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

# ── Credentials ───────────────────────────────────────────────────────────────
load_dotenv()
USERNAME = os.getenv("INSTAGRAM_USERNAME")
PASSWORD = os.getenv("INSTAGRAM_PASSWORD")

if not USERNAME or not PASSWORD:
    raise ValueError("Missing INSTAGRAM_USERNAME or INSTAGRAM_PASSWORD in .env file")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "data", "raw", "instagram"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Config ────────────────────────────────────────────────────────────────────
post_shortcodes = [
    "DXN9Avbku-x",
    "DUj_OTXD8Xv",
    "DRBtPUMjxho",
]
MAX_COMMENTS_PER_POST = 200

# ── Helpers ───────────────────────────────────────────────────────────────────
def human_delay(min_sec=1.5, max_sec=3.5):
    time.sleep(random.uniform(min_sec, max_sec))


def extract_shortcode(url_or_code: str) -> str:
    match = re.search(r"instagram\.com/(?:p|reel)/([A-Za-z0-9_-]+)", url_or_code)
    return match.group(1) if match else url_or_code.strip().strip("/").replace("p/", "")


def try_dismiss_popups(driver):
    popup_xpaths = [
        "//button[contains(text(),'Allow')]",
        "//button[contains(text(),'Accept')]",
        "//button[contains(text(),'Only allow essential cookies')]",
        "//button[@title='Allow all cookies']",
        "//button[text()='Not now']",
        "//button[text()='Not Now']",
        "//button[text()='Save Info']",
    ]
    for xpath in popup_xpaths:
        try:
            btn = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
            btn.click()
            human_delay(0.5, 1.0)
        except Exception:
            pass


def init_driver():
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless=new")  # Uncomment for headless
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-infobars")
    options.add_argument("--start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    return driver


def login(driver):
    print("Opening Instagram login page...")
    driver.get("https://www.instagram.com/")
    human_delay(3, 5)
    try_dismiss_popups(driver)
    human_delay(1, 2)

    driver.get("https://www.instagram.com/accounts/login/")
    human_delay(3, 5)
    try_dismiss_popups(driver)

    user_input = None
    for by, selector in [
        (By.NAME,         "username"),
        (By.CSS_SELECTOR, "input[name='username']"),
        (By.XPATH,        "//input[@name='username']"),
    ]:
        try:
            user_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((by, selector))
            )
            break
        except TimeoutException:
            continue

    if user_input is None:
        print("Could not find login form. Please log in manually.")
        input("Press Enter once fully logged in...")
        return

    user_input.click()
    user_input.clear()
    for char in USERNAME:
        user_input.send_keys(char)
        time.sleep(random.uniform(0.05, 0.18))
    human_delay(0.5, 1.2)

    pass_input = None
    for by, selector in [
        (By.NAME,         "password"),
        (By.CSS_SELECTOR, "input[type='password']"),
        (By.XPATH,        "//input[@name='password']"),
    ]:
        try:
            pass_input = driver.find_element(by, selector)
            break
        except NoSuchElementException:
            continue

    if pass_input is None:
        print("Could not find password field.")
        input("Press Enter once fully logged in...")
        return

    pass_input.click()
    pass_input.clear()
    for char in PASSWORD:
        pass_input.send_keys(char)
        time.sleep(random.uniform(0.05, 0.18))
    human_delay(0.8, 1.5)
    pass_input.send_keys(Keys.RETURN)
    print("Submitted login, waiting for home page...")

    try:
        WebDriverWait(driver, 20).until(
            EC.any_of(
                EC.presence_of_element_located((By.XPATH, "//nav")),
                EC.url_contains("instagram.com/?"),
            )
        )
        print(f"✅ Logged in as: {USERNAME}")
    except TimeoutException:
        print("Needs verification. Complete it in the browser.")
        input("Press Enter once fully logged in...")

    human_delay(2, 3)
    try_dismiss_popups(driver)
    human_delay(1, 2)


# ── Debug helper ──────────────────────────────────────────────────────────────
def debug_page_structure(driver):
    """Prints what elements are found — paste output here to fix selectors."""
    print("\n===== DEBUG: Page element scan =====")
    tests = [
        ("ul li",               "List items"),
        ("span[dir='auto']",    "span[dir=auto]  ← comment text lives here"),
        ("time",                "Time elements"),
        ("a[role='link']",      "Role=link anchors (usernames)"),
        ("div[role='button']",  "Clickable divs"),
        ("article",             "Article tags"),
    ]
    for sel, label in tests:
        els = driver.find_elements(By.CSS_SELECTOR, sel)
        print(f"  [{len(els):3d}] {label}")

    print("\n  First 30 span[dir='auto'] texts:")
    spans = driver.find_elements(By.CSS_SELECTOR, "span[dir='auto']")
    for i, s in enumerate(spans[:30]):
        txt = s.text.strip()
        if txt:
            print(f"    [{i:02d}] {repr(txt[:120])}")
    print("===== END DEBUG =====\n")


# ── Core selectors (stable across Instagram redesigns) ────────────────────────
UI_SKIP = {
    "Like", "Reply", "Follow", "Following", "Message", "Share", "Save",
    "More", "Loading", "View", "replies", "comments", "likes", "Verified",
    "Suggested", "Hide", "Report", "Translate", "See translation",
    "See less", "Load more", "View more",
}


def get_username_near(span):
    try:
        el = span
        for _ in range(7):
            el = el.find_element(By.XPATH, "..")
            links = el.find_elements(By.XPATH, ".//a[@href]")
            for link in links:
                href = link.get_attribute("href") or ""
                text = link.text.strip()
                if re.match(r"https://www\.instagram\.com/[^/?#]+/?$", href) and text:
                    return text
    except Exception:
        pass
    return ""


def get_timestamp_near(span):
    try:
        el = span
        for _ in range(7):
            el = el.find_element(By.XPATH, "..")
            times = el.find_elements(By.TAG_NAME, "time")
            if times:
                return times[0].get_attribute("datetime") or times[0].text.strip()
    except Exception:
        pass
    return ""


def click_buttons_by_text(driver, keywords):
    """Click all visible buttons/spans whose text contains any of the keywords."""
    clicked = 0
    xpath = " or ".join([f"contains(text(),'{k}')" for k in keywords])
    for tag in ["span", "button", "div"]:
        try:
            els = driver.find_elements(By.XPATH, f"//{tag}[{xpath}]")
            for el in els:
                try:
                    driver.execute_script("arguments[0].scrollIntoView(true);", el)
                    driver.execute_script("arguments[0].click();", el)
                    clicked += 1
                    time.sleep(random.uniform(0.4, 0.9))
                except Exception:
                    pass
        except Exception:
            pass
    return clicked


def scrape_comments(driver, shortcode, max_comments=200, debug=False):
    sc  = extract_shortcode(shortcode)
    url = f"https://www.instagram.com/p/{sc}/"
    print(f"\n📌 Opening: {url}")
    driver.get(url)
    human_delay(4, 6)
    try_dismiss_popups(driver)
    human_delay(1, 2)

    # Scroll down to load comments section
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 3);")
    human_delay(2, 3)

    if debug:
        debug_page_structure(driver)

    # Get post owner
    post_owner = ""
    try:
        owner_el   = driver.find_element(
            By.XPATH,
            "//header//a[not(contains(@href,'/p/')) and not(contains(@href,'/explore/'))]"
        )
        post_owner = owner_el.text.strip()
    except Exception:
        pass

    comments_data = []
    seen_texts    = set()
    no_new_count  = 0

    while len(comments_data) < max_comments:
        prev_count = len(comments_data)

        # Expand reply threads
        r = click_buttons_by_text(driver, ["View replies", "View reply"])
        if r:
            human_delay(1.0, 2.0)

        # ── PRIMARY METHOD: span[dir='auto'] ──────────────────────────────
        # Instagram renders ALL comment text inside span[dir='auto']
        # This selector survives class name obfuscation.
        spans = driver.find_elements(By.CSS_SELECTOR, "span[dir='auto']")

        for span in spans:
            try:
                text = span.text.strip()
            except Exception:
                continue

            if not text or len(text) < 2 or text in seen_texts or text in UI_SKIP:
                continue
            # Skip pure emoji-only or single-word UI tokens under 3 chars
            if len(text) <= 2 and not any(c.isalpha() for c in text):
                continue

            seen_texts.add(text)
            username  = get_username_near(span)
            timestamp = get_timestamp_near(span)

            comments_data.append({
                "post_id"       : sc,
                "post_url"      : url,
                "post_owner"    : post_owner,
                "comment_type"  : "comment",
                "parent_comment": "",
                "username"      : username,
                "comment_text"  : text,
                "timestamp"     : timestamp,
                "platform"      : "instagram",
                "label"         : "",
                "severity"      : "",
                "notes"         : ""
            })

        # Load more comments
        loaded = click_buttons_by_text(
            driver, ["Load more comments", "View more comments"]
        )
        if not loaded:
            driver.execute_script("window.scrollBy(0, 800);")
            human_delay(1.5, 2.5)
        else:
            human_delay(2.5, 4.0)

        new_count = len(comments_data)
        if new_count == prev_count:
            no_new_count += 1
            if no_new_count >= 5:
                print("  No more new comments found.")
                break
        else:
            no_new_count = 0
            print(f"  Collected so far: {new_count}")

    print(f"  ✅ Total from {sc}: {len(comments_data)}")
    return comments_data


# ── Save to Excel ─────────────────────────────────────────────────────────────
def save_to_excel(df, output_path):
    wb     = openpyxl.Workbook()
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cmt_fill  = PatternFill("solid", start_color="EBF3FB")
    sum_fill  = PatternFill("solid", start_color="DEEAF1")

    # Sheet 1 — Comments Data
    ws1 = wb.active
    ws1.title = "Comments Data"

    columns = [
        ("Post ID",14),("Post URL",40),("Post Owner",18),("Type",10),
        ("Parent Comment",45),("Username",20),("Comment Text",60),
        ("Timestamp",22),("Platform",12),("Label",15),("Severity",12),("Notes",30),
    ]
    for ci, (name, width) in enumerate(columns, 1):
        c = ws1.cell(row=1, column=ci, value=name)
        c.font=hdr_font; c.fill=hdr_fill; c.alignment=center; c.border=border
        ws1.column_dimensions[get_column_letter(ci)].width = width

    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "A2"

    keys = ["post_id","post_url","post_owner","comment_type","parent_comment",
            "username","comment_text","timestamp","platform","label","severity","notes"]

    for ri, row in df.iterrows():
        for ci, key in enumerate(keys, 1):
            c = ws1.cell(row=ri+2, column=ci, value=row.get(key,""))
            c.fill=cmt_fill; c.border=border
            c.font=Font(name="Arial",size=10); c.alignment=left

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Sheet 2 — Summary
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 30

    def wrow(s, r, label, value, hdr=False):
        for ci, v in enumerate([label, value], 1):
            c = s.cell(row=r, column=ci, value=v)
            c.border=border; c.alignment=left
            c.font=Font(name="Arial",bold=hdr,color="FFFFFF" if hdr else "000000",size=11)
            c.fill=hdr_fill if hdr else sum_fill

    wrow(ws2,1,"Field","Value",hdr=True)
    wrow(ws2,2,"Scrape Date",     datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    wrow(ws2,3,"Total Posts",     len(post_shortcodes))
    wrow(ws2,4,"Total Comments",  len(df))
    wrow(ws2,5,"Unique Users",    df["username"].nunique())
    wrow(ws2,6,"Platform",        "Instagram")
    wrow(ws2,7,"Post Shortcodes", " | ".join([extract_shortcode(s) for s in post_shortcodes]))

    # Sheet 3 — Column Guide
    ws3 = wb.create_sheet("Column Guide")
    ws3.column_dimensions["A"].width=22
    ws3.column_dimensions["B"].width=55
    ws3.column_dimensions["C"].width=32

    guide = [
        ("Column","Description","Example / Values",True),
        ("Post ID",      "Instagram shortcode",               "DXN9Avbku-x",                   False),
        ("Post URL",     "Full link to the post",             "https://instagram.com/p/…",     False),
        ("Post Owner",   "Username of page that posted",      "memezar",                       False),
        ("Type",         "comment or reply",                  "comment / reply",               False),
        ("Parent Comment","Comment being replied to",         "lol this is so true…",          False),
        ("Username",     "Who wrote the comment",             "user123",                       False),
        ("Comment Text", "Full text of comment or reply",    "Haha 😂",                       False),
        ("Timestamp",    "Date/time posted",                  "2024-01-15T10:30:00.000Z",      False),
        ("Platform",     "Source platform",                   "instagram",                     False),
        ("Label",        "Analyst: sentiment",                "positive / negative / neutral", False),
        ("Severity",     "Analyst: toxicity level",           "low / medium / high",           False),
        ("Notes",        "Analyst notes",                     "Contains Tamil slang",          False),
    ]
    for ri, (col,desc,ex,hdr) in enumerate(guide, 1):
        for ci, v in enumerate([col,desc,ex], 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.border=border; c.alignment=left
            c.font=Font(name="Arial",bold=hdr,color="FFFFFF" if hdr else "000000",size=10)
            c.fill=hdr_fill if hdr else cmt_fill

    wb.save(output_path)
    print(f"\n✅ Excel saved → {output_path}")
    print(f"   📊 Sheets: 'Comments Data' | 'Summary' | 'Column Guide'")


# ── Main ──────────────────────────────────────────────────────────────────────
# DEBUG_MODE=True prints what elements exist on the page.
# After first run paste the ===== DEBUG ===== output here if still 0 comments.
DEBUG_MODE = True

driver       = init_driver()
all_comments = []

try:
    login(driver)
    human_delay(2, 3)

    for sc in post_shortcodes:
        try:
            comments = scrape_comments(driver, sc, MAX_COMMENTS_PER_POST, debug=DEBUG_MODE)
            all_comments.extend(comments)
            DEBUG_MODE = False  # debug only first post
        except Exception as e:
            print(f"Skipping {sc}: {e}")

finally:
    driver.quit()

# ── Save ──────────────────────────────────────────────────────────────────────
if all_comments:
    df = pd.DataFrame(all_comments)
    df.drop_duplicates(subset=["comment_text"], inplace=True)
    df = df[df["comment_text"].str.strip() != ""]
    df.reset_index(drop=True, inplace=True)

    filename    = datetime.now().strftime("%Y%m%d_%H%M%S") + "_instagram_comments.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    save_to_excel(df, output_path)
    print(f"\n📊 Total rows saved: {len(df)}")
else:
    print("\n⚠️  No comments scraped.")
    print("    Paste the ===== DEBUG ===== section from your terminal here")
    print("    so the selectors can be fixed for your Instagram version.")