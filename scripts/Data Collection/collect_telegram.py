"""
Telegram Message Scraper using Telethon
Scrapes public channels/groups — extracts:
  • Text messages / comments  → Sheet: "Text Messages"
  • Meme/image OCR text       → Sheet: "Image OCR Text"
Images are processed in memory only — nothing saved to disk.

Install:
    pip install telethon pandas openpyxl python-dotenv pillow pytesseract

System requirement (Tesseract OCR engine):
    sudo apt install tesseract-ocr tesseract-ocr-tam tesseract-ocr-hin tesseract-ocr-eng

Usage:
    python scrape_telegram.py
"""

import os
import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from telethon.sync import TelegramClient
from telethon.errors import ChannelPrivateError, UsernameNotOccupiedError
from telethon.tl.types import MessageMediaPhoto, MessageMediaDocument
from PIL import Image
import pytesseract

load_dotenv()

API_ID   = int(os.getenv("TELEGRAM_API_ID"))
API_HASH = os.getenv("TELEGRAM_API_HASH")

if not API_ID or not API_HASH:
    raise ValueError("Missing TELEGRAM_API_ID or TELEGRAM_API_HASH in .env file")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.abspath(os.path.join(BASE_DIR, "..", "data", "raw", "telegram"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

SESSION_FILE = os.path.join(BASE_DIR, "safevoice_session")

# OCR language config — add/remove based on your channels
# Options: eng, tam (Tamil), hin (Hindi), mal (Malayalam), tel (Telugu)
OCR_LANG = "eng+tam+hin"

# ── Channels to scrape ───────────────────────────────────────────────────────
public_channels = [
    "tamilcomedymemes98",
]

LIMIT_PER_CHANNEL = 1000


# ── Helpers ──────────────────────────────────────────────────────────────────

def ocr_from_bytes(image_bytes: bytes) -> str:
    """Extract text from image bytes in memory. Image is never saved to disk."""
    try:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        return pytesseract.image_to_string(image, lang=OCR_LANG).strip()
    except Exception as e:
        print(f"    [OCR error] {e}")
        return ""


def is_image_media(msg) -> bool:
    if isinstance(msg.media, MessageMediaPhoto):
        return True
    if isinstance(msg.media, MessageMediaDocument):
        doc = msg.media.document
        if doc and doc.mime_type and doc.mime_type.startswith("image/"):
            return True
    return False


def style_header_row(ws, headers: list, header_fill: str):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color=header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30


def style_data_rows(ws, num_rows: int, num_cols: int):
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_alt = PatternFill("solid", start_color="EBF3FB")
    for row in range(2, num_rows + 2):
        fill = fill_alt if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if fill:
                cell.fill = fill


def write_sheet(wb, sheet_name: str, df: pd.DataFrame, header_fill: str, col_widths: dict):
    ws = wb.create_sheet(title=sheet_name)
    headers = list(df.columns)
    style_header_row(ws, headers, header_fill)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
    style_data_rows(ws, len(df), len(headers))
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Scrape ───────────────────────────────────────────────────────────────────
text_messages  = []
image_messages = []

with TelegramClient(SESSION_FILE, API_ID, API_HASH) as client:
    for channel in public_channels:
        print(f"\nScraping channel: @{channel}")
        try:
            for msg in client.iter_messages(channel, limit=LIMIT_PER_CHANNEL):

                caption = (msg.text or "").strip()

                # ── Image message — OCR in memory, no file saved ─────────────
                if is_image_media(msg):
                    try:
                        image_bytes = client.download_media(msg, bytes)  # to memory only
                        if image_bytes:
                            ocr_text = ocr_from_bytes(image_bytes)
                            del image_bytes  # free memory immediately after OCR

                            # Only keep rows where OCR actually found text
                            if ocr_text and len(ocr_text.strip()) > 3:
                                image_messages.append({
                                    "OCR Text":   ocr_text,
                                    "Caption":    caption,
                                    "Channel":    channel,
                                    "Message ID": msg.id,
                                    "Sender ID":  msg.sender_id,
                                    "Timestamp":  str(msg.date),
                                    "Is Reply":   msg.is_reply,
                                    "Label":      "",
                                    "Severity":   "",
                                })
                                print(f"    [image] msg_id={msg.id} | OCR chars={len(ocr_text)}")
                    except Exception as e:
                        print(f"    [image error] msg_id={msg.id}: {e}")

                # ── Plain text / comment message ─────────────────────────────
                elif caption and len(caption) > 3:
                    text_messages.append({
                        "Text":       caption,
                        "Channel":    channel,
                        "Message ID": msg.id,
                        "Sender ID":  msg.sender_id,
                        "Timestamp":  str(msg.date),
                        "Is Reply":   msg.is_reply,
                        "Label":      "",
                        "Severity":   "",
                    })

            t = sum(1 for m in text_messages  if m["Channel"] == channel)
            i = sum(1 for m in image_messages if m["Channel"] == channel)
            print(f"  → {t} text msgs | {i} image OCR msgs")

        except ChannelPrivateError:
            print(f"  Skipping @{channel}: private channel")
        except UsernameNotOccupiedError:
            print(f"  Skipping @{channel}: channel not found")
        except Exception as e:
            print(f"  Skipping @{channel}: {e}")


# ── Build Excel ──────────────────────────────────────────────────────────────
if text_messages or image_messages:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Text Messages ───────────────────────────────────────────────
    if text_messages:
        df_text = pd.DataFrame(text_messages)
        df_text.drop_duplicates(subset=["Text"], inplace=True)
        df_text = df_text[df_text["Text"].str.strip() != ""]

        write_sheet(
            wb, "Text Messages", df_text,
            header_fill="1F4E79",   # dark blue
            col_widths={"A": 65, "B": 20, "C": 12, "D": 14, "E": 22, "F": 10, "G": 15, "H": 15}
        )
        print(f"\n📝 Text Messages sheet: {len(df_text)} rows")

    # ── Sheet 2: Image OCR Text ──────────────────────────────────────────────
    if image_messages:
        df_img = pd.DataFrame(image_messages)
        df_img.drop_duplicates(subset=["OCR Text"], inplace=True)
        df_img = df_img[df_img["OCR Text"].str.strip() != ""]

        write_sheet(
            wb, "Image OCR Text", df_img,
            header_fill="375623",   # dark green
            col_widths={"A": 65, "B": 40, "C": 18, "D": 12, "E": 14, "F": 22, "G": 10, "H": 15, "I": 15}
        )
        print(f"🖼️  Image OCR Text sheet: {len(df_img)} rows")

    # ── Save Excel ───────────────────────────────────────────────────────────
    filename    = datetime.now().strftime("%Y%m%d_%H%M%S") + "_telegram_scraped.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(output_path)

    print(f"\n✅ Excel saved to:\n{output_path}")
    print(f"   Sheets: 'Text Messages' | 'Image OCR Text'")

else:
    print("\nNo messages scraped. Check your channel names.")